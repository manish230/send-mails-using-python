# import openpyxl module 
import openpyxl 

import smtplib

from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

gmail_user = "abc@gmail.com"
gmail_appPassword = "Learningbud@12"
# Give the location of the file on local machine
path = "users.xlsx"
  
# To open the workbook  
# workbook object is created 
wb_obj = openpyxl.load_workbook(path) 
  
# Get workbook active sheet object 
# from the active attribute 
sheet = wb_obj.active 
m_row = sheet.max_row
max_column=sheet.max_column



for i in range(1,m_row+1):
     text = 'Hi \nYour  username, password and mail ID have been shared. Please find below credentials ' + '\n'
     # iterate over all columns
     for j in range(1,max_column+1):
          cell_obj=sheet.cell(row=i,column=j)
          subject = 'Credentials'
          sent_from = ['abc@gmail.com']
          message = MIMEMultipart("alternative")
          message["Subject"] = "credentials"

          message["From"] = str(sent_from)

          text += str(cell_obj.value) + '\n'
          if (j == 3):
            print('hello')  
            to = str(cell_obj.value)
            message["To"] = str(to)
            msg = MIMEText(text, "plain")
            message.attach(msg)
            server = smtplib.SMTP_SSL('smtp.gmail.com', 465)
            server.login(gmail_user, gmail_appPassword)
            server.sendmail(sent_from, to, message.as_string())

            server.quit()
